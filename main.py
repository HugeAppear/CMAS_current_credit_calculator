import openpyxl
import datetime


wb = openpyxl.load_workbook('CMAS tracker.xlsx')
ws = wb.active
# ws = wb.worksheets[0]
ws['C1'] = 'CMAS eligibility date'
ws['A3'] = '최유창'
ws['B3'] = datetime.date(2020, 12, 31)
ws['D1'] = 'CMAS credit'

ref_wb = openpyxl.load_workbook('CMAS reference.xlsx')
ref_ws = ref_wb.worksheets[0]

def cal_CMAS_credit(col_num):
    delta_181 = datetime.timedelta(days=181)

    CMAS_elig_date = ws['B'+str(col_num)].value + delta_181

    ws['C'+str(col_num)].value = CMAS_elig_date.strftime('%Y-%m-%d')
    day = CMAS_elig_date.day


    if CMAS_elig_date.month in [1, 3, 5, 7, 8, 10, 12]:
        cell = 'E' + str(day + 2)
        first_month_credit = ref_ws[cell].value
    elif CMAS_elig_date.month in [4, 6, 9, 11]:
        cell = 'D' + str(day + 2)
        first_month_credit = ref_ws[cell].value
    else: # 2월이면
        if CMAS_elig_date.year % 4 == 0: # 윤년이면
            cell = 'B' + str(day + 2)
            first_month_credit = ref_ws[cell].value
        else: # 윤년아니면
            cell = 'C' + str(day + 2)
            first_month_credit = ref_ws[cell].value

    credit = first_month_credit

    current_year, current_month = datetime.datetime.today().year, datetime.datetime.today().month
    elig_year, elig_month = CMAS_elig_date.year, CMAS_elig_date.month

    if current_year == elig_year: # 만약 같은 년도라면
        # ex) current가 21-12, elig가 21-04라면, 12-4=8개월 쌓여야함!
        month_delta = current_month - elig_month
        # m
    elif current_year == elig_year+1:# 만약 한 해 뒤라면
        # ex) 22-2와 21-4 라면, 2-4+12 = 10
        month_delta = current_month - elig_month + 12
    else: # 만약 두 해 뒤라면
        month_delta = current_month - elig_month + 24


    credit += month_delta*7.26
    ws['D'+str(col_num)] = '$' + str(credit)


cal_CMAS_credit(2) # 2는 김태현
cal_CMAS_credit(3)

# 엑셀 파일 저장하기
wb.save('CMAS tracker.xlsx')









